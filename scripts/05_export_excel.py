from pathlib import Path
"""
IAQF 2026 — Excel Export Script v2 (fast, using xlsxwriter)
"""

import pandas as pd
import numpy as np
import os

OUT_DIR    = str(Path(__file__).parent.parent / "data" / "parquet")
EXCEL_PATH = os.path.join(OUT_DIR, "IAQF_DataFinal.xlsx")

# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading feature panels...")
df_1m  = pd.read_parquet(os.path.join(OUT_DIR, "iaqf_features_1m.parquet"))
df_1h  = pd.read_parquet(os.path.join(OUT_DIR, "iaqf_features_1h.parquet"))
df_day = pd.read_parquet(os.path.join(OUT_DIR, "iaqf_daily_summary.parquet"))

# Convert ALL timezone-aware datetime columns to timezone-naive for Excel
def strip_tz(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                df[col] = df[col].dt.tz_localize(None)
            except TypeError:
                df[col] = df[col].dt.tz_convert(None)
    return df

df_1m  = strip_tz(df_1m)
df_1h  = strip_tz(df_1h)
df_day = strip_tz(df_day)

# ── Key column selections ─────────────────────────────────────────────────────
cols_1m = [
    "timestamp_utc", "regime", "day_of_study", "hour_utc",
    # Binance.US BTC prices
    "bnus_btcusd_open", "bnus_btcusd_high", "bnus_btcusd_low", "bnus_btcusd_close", "bnus_btcusd_vol",
    "bnus_btcusdt_open", "bnus_btcusdt_high", "bnus_btcusdt_low", "bnus_btcusdt_close", "bnus_btcusdt_vol",
    "bnus_btcusdc_open", "bnus_btcusdc_high", "bnus_btcusdc_low", "bnus_btcusdc_close", "bnus_btcusdc_vol",
    "bnus_btcbusd_open", "bnus_btcbusd_high", "bnus_btcbusd_low", "bnus_btcbusd_close", "bnus_btcbusd_vol",
    # Coinbase BTC prices
    "cb_btcusd_open", "cb_btcusd_high", "cb_btcusd_low", "cb_btcusd_close", "cb_btcusd_vol",
    "cb_btcusdt_open", "cb_btcusdt_high", "cb_btcusdt_low", "cb_btcusdt_close", "cb_btcusdt_vol",
    # Stablecoin FX
    "usdt_usd_close", "usdc_usd_close", "usdc_usdt_close", "busd_usdt_close",
    # Mid-prices
    "mid_bnus_btcusd", "mid_bnus_btcusdt", "mid_bnus_btcusdc", "mid_bnus_btcbusd",
    "mid_cb_btcusd", "mid_cb_btcusdt",
    "mid_usdt_usd", "mid_usdc_usd",
    # Spreads
    "spread_bnus_btcusd", "spread_bnus_btcusdt", "spread_bnus_btcusdc",
    "spread_cb_btcusd", "spread_cb_btcusdt",
    "rel_spread_bnus_btcusd", "rel_spread_bnus_btcusdt", "rel_spread_bnus_btcusdc",
    # LOP deviations
    "lop_bnus_usdt_vs_usd", "lop_bnus_usdc_vs_usd",
    "lop_bnus_usdt_vs_busd", "lop_bnus_usdc_vs_busd",
    "lop_cb_usdt_vs_usd", "lop_bnus_vs_cb_usd",
    "abs_lop_bnus_usdt_vs_usd", "abs_lop_bnus_usdc_vs_usd",
    "abs_lop_cb_usdt_vs_usd",
    # Stablecoin FX deviations
    "log_usdt_usd_dev", "log_usdc_usd_dev", "log_usdc_usdt", "log_busd_usdt",
    "stablecoin_fx_usdt", "stablecoin_fx_usdc",
    "lop_residual_bnus_usdt", "lop_residual_bnus_usdc",
    # Realized vol & returns
    "rv60_bnus_btcusd", "rv60_bnus_btcusdt", "rv60_bnus_btcusdc",
    "logret_bnus_btcusd", "logret_bnus_btcusdt", "logret_bnus_btcusdc",
    # Volume metrics
    "total_vol_bnus_btc", "vol_share_usdt_bnus", "vol_share_usdc_bnus",
    "vol_share_usd_bnus", "vol_share_busd_bnus",
    "bnus_btcusd_trades", "bnus_btcusdt_trades", "bnus_btcusdc_trades",
]
cols_1m = [c for c in cols_1m if c in df_1m.columns]

# ── Regime summary ────────────────────────────────────────────────────────────
df_1m_bps = df_1m.copy()
df_1m_bps["abs_lop_bnus_usdt_vs_usd_bps"] = df_1m_bps["abs_lop_bnus_usdt_vs_usd"] * 10000
df_1m_bps["abs_lop_bnus_usdc_vs_usd_bps"] = df_1m_bps["abs_lop_bnus_usdc_vs_usd"] * 10000
df_1m_bps["rel_spread_bnus_btcusd_bps"]   = df_1m_bps["rel_spread_bnus_btcusd"]   * 10000
df_1m_bps["rel_spread_bnus_btcusdt_bps"]  = df_1m_bps["rel_spread_bnus_btcusdt"]  * 10000
df_1m_bps["rel_spread_bnus_btcusdc_bps"]  = df_1m_bps["rel_spread_bnus_btcusdc"]  * 10000

df_regime = df_1m_bps.groupby("regime").agg(
    n_minutes=("timestamp_utc", "count"),
    btcusd_mean=("bnus_btcusd_close", "mean"),
    btcusd_min=("bnus_btcusd_close", "min"),
    btcusd_max=("bnus_btcusd_close", "max"),
    usdc_usd_mean=("usdc_usd_close", "mean"),
    usdc_usd_min=("usdc_usd_close", "min"),
    usdc_usd_max=("usdc_usd_close", "max"),
    usdt_usd_mean=("usdt_usd_close", "mean"),
    usdt_usd_min=("usdt_usd_close", "min"),
    usdt_usd_max=("usdt_usd_close", "max"),
    lop_usdt_mean_bps=("abs_lop_bnus_usdt_vs_usd_bps", "mean"),
    lop_usdt_max_bps=("abs_lop_bnus_usdt_vs_usd_bps", "max"),
    lop_usdc_mean_bps=("abs_lop_bnus_usdc_vs_usd_bps", "mean"),
    lop_usdc_max_bps=("abs_lop_bnus_usdc_vs_usd_bps", "max"),
    rel_spread_usd_bps=("rel_spread_bnus_btcusd_bps", "mean"),
    rel_spread_usdt_bps=("rel_spread_bnus_btcusdt_bps", "mean"),
    rel_spread_usdc_bps=("rel_spread_bnus_btcusdc_bps", "mean"),
    rv60_btcusd_mean=("rv60_bnus_btcusd", "mean"),
    rv60_btcusdt_mean=("rv60_bnus_btcusdt", "mean"),
    btcusd_vol_total=("bnus_btcusd_vol", "sum"),
    btcusdt_vol_total=("bnus_btcusdt_vol", "sum"),
    btcusdc_vol_total=("bnus_btcusdc_vol", "sum"),
).reset_index()

regime_order = {"pre_crisis": 0, "crisis": 1, "recovery": 2, "post": 3}
df_regime["_ord"] = df_regime["regime"].map(regime_order)
df_regime = df_regime.sort_values("_ord").drop(columns=["_ord"])

# ── LOP Analysis sheet ────────────────────────────────────────────────────────
cols_lop = [
    "timestamp_utc", "regime", "day_of_study",
    "bnus_btcusd_close", "bnus_btcusdt_close", "bnus_btcusdc_close", "bnus_btcbusd_close",
    "cb_btcusd_close", "cb_btcusdt_close",
    "mid_bnus_btcusd", "mid_bnus_btcusdt", "mid_bnus_btcusdc",
    "lop_bnus_usdt_vs_usd", "lop_bnus_usdc_vs_usd",
    "lop_bnus_usdt_vs_busd", "lop_bnus_usdc_vs_busd",
    "lop_cb_usdt_vs_usd", "lop_bnus_vs_cb_usd",
    "abs_lop_bnus_usdt_vs_usd", "abs_lop_bnus_usdc_vs_usd",
    "stablecoin_fx_usdt", "stablecoin_fx_usdc",
    "lop_residual_bnus_usdt", "lop_residual_bnus_usdc",
    "spread_bnus_btcusd", "spread_bnus_btcusdt", "spread_bnus_btcusdc",
    "rel_spread_bnus_btcusd", "rel_spread_bnus_btcusdt", "rel_spread_bnus_btcusdc",
    "rv60_bnus_btcusd", "rv60_bnus_btcusdt",
]
cols_lop = [c for c in cols_lop if c in df_1m.columns]

# ── Stablecoin FX sheet ───────────────────────────────────────────────────────
cols_fx = [
    "timestamp_utc", "regime", "day_of_study",
    "usdt_usd_close", "usdt_usd_high", "usdt_usd_low", "usdt_usd_vol",
    "usdc_usd_close", "usdc_usd_high", "usdc_usd_low", "usdc_usd_vol",
    "usdc_usdt_close", "usdc_usdt_vol",
    "busd_usdt_close", "busd_usdt_vol",
    "log_usdt_usd_dev", "log_usdc_usd_dev",
    "log_usdc_usdt", "log_busd_usdt",
    "stablecoin_fx_usdt", "stablecoin_fx_usdc",
    "mid_usdt_usd", "mid_usdc_usd",
]
cols_fx = [c for c in cols_fx if c in df_1m.columns]

# ── Data Inventory ────────────────────────────────────────────────────────────
df_inv = pd.DataFrame({
    "Source": [
        "Binance.US", "Binance.US", "Binance.US", "Binance.US",
        "Binance.US", "Binance.US", "Binance.US", "Binance.US",
        "Coinbase", "Coinbase", "Coinbase",
    ],
    "Trading_Pair": [
        "BTC/USDT", "BTC/USDC", "BTC/USD", "BTC/BUSD",
        "USDT/USD", "USDC/USD", "USDC/USDT", "BUSD/USDT",
        "BTC-USD", "BTC-USDC", "BTC-USDT",
    ],
    "API_Endpoint": [
        "api.binance.us/api/v3/klines?symbol=BTCUSDT",
        "api.binance.us/api/v3/klines?symbol=BTCUSDC",
        "api.binance.us/api/v3/klines?symbol=BTCUSD",
        "api.binance.us/api/v3/klines?symbol=BTCBUSD",
        "api.binance.us/api/v3/klines?symbol=USDTUSD",
        "api.binance.us/api/v3/klines?symbol=USDCUSD",
        "api.binance.us/api/v3/klines?symbol=USDCUSDT",
        "api.binance.us/api/v3/klines?symbol=BUSDUSDT",
        "api.coinbase.com/.../BTC-USD/candles",
        "api.coinbase.com/.../BTC-USDC/candles",
        "api.coinbase.com/.../BTC-USDT/candles",
    ],
    "Rows": [30240, 30240, 30240, 30240, 30240, 30240, 30240, 30240, 29950, 29950, 24478],
    "Frequency": ["1-minute"] * 11,
    "Null_Pct_Close": ["0.00%"] * 8 + ["0.90%", "0.90%", "1.07%"],
    "Fields_Available": [
        "open, high, low, close, volume, num_trades, quote_vol",
        "open, high, low, close, volume, num_trades, quote_vol",
        "open, high, low, close, volume, num_trades, quote_vol",
        "open, high, low, close, volume, num_trades, quote_vol",
        "open, high, low, close, volume",
        "open, high, low, close, volume",
        "close, volume",
        "close, volume",
        "open, high, low, close, volume",
        "open, high, low, close, volume (mirrors BTC-USD)",
        "open, high, low, close, volume",
    ],
    "Notes": [
        "Primary USDT-quoted BTC pair",
        "Primary USDC-quoted BTC pair",
        "Direct fiat USD pair — primary USD reference",
        "BUSD-quoted BTC (BUSD≈USD, used as USD proxy)",
        "USDT/USD spot rate (stablecoin FX)",
        "USDC/USD spot rate (stablecoin FX, shows de-peg)",
        "USDC/USDT cross rate",
        "BUSD/USDT cross rate",
        "Primary USD-quoted BTC (Coinbase)",
        "USDC-quoted BTC; public candle API mirrors BTC-USD",
        "USDT-quoted BTC (Coinbase); lower liquidity",
    ],
})

# ── README content ────────────────────────────────────────────────────────────
readme_rows = [
    ["IAQF Student Competition 2026 — Data Package"],
    ["Cross-Currency Dynamics in Cryptocurrencies under Stablecoin Regulation"],
    [""],
    ["STUDY OVERVIEW"],
    ["This dataset supports the analysis of cross-currency price dynamics in BTC spot markets"],
    ["during the USDC de-peg event of March 10-13, 2023, caused by Silicon Valley Bank's failure."],
    [""],
    ["DATA SOURCES"],
    ["  - Binance.US (api.binance.us) — 1-minute OHLCV klines via REST API"],
    ["  - Coinbase Advanced Trade (api.coinbase.com) — 1-minute candles via REST API"],
    [""],
    ["TIME PERIOD: March 1, 2023 00:00:00 UTC to March 21, 2023 23:59:00 UTC"],
    ["FREQUENCY: 1-minute OHLCV candles (30,240 rows per pair)"],
    [""],
    ["EVENT REGIMES"],
    ["  pre_crisis  : March 1-9, 2023   (baseline)"],
    ["  crisis      : March 10-12, 2023 (SVB collapse → USDC reserve freeze → de-peg)"],
    ["  recovery    : March 13-15, 2023 (Fed/FDIC backstop → Circle resumes redemptions)"],
    ["  post        : March 16-21, 2023 (USDC re-pegged, market normalizes)"],
    [""],
    ["KEY COMPUTED VARIABLES"],
    ["  mid_*              : Mid-price = (High + Low) / 2"],
    ["  spread_*           : Intrabar spread = High - Low (Parkinson liquidity proxy)"],
    ["  rel_spread_*       : Relative spread = spread / mid-price"],
    ["  lop_*              : Log LOP deviation = log(BTC/X) - log(BTC/USD)"],
    ["  abs_lop_*          : |LOP deviation| (absolute value)"],
    ["  log_usdt_usd_dev   : log(USDT/USD) — deviation from parity"],
    ["  log_usdc_usd_dev   : log(USDC/USD) — deviation from parity"],
    ["  stablecoin_fx_usdt : -log(USDT/USD) — FX component of USDT basis"],
    ["  stablecoin_fx_usdc : -log(USDC/USD) — FX component of USDC basis"],
    ["  lop_residual_*     : LOP deviation minus stablecoin FX component"],
    ["  rv60_*             : 60-min rolling realized volatility (sum sq. log-returns)"],
    ["  logret_*           : 1-minute log return"],
    ["  vol_share_*        : Volume share by quote currency (Binance.US)"],
    ["  regime             : Event regime label"],
    [""],
    ["SHEETS"],
    ["  1min_Panel     : Full 1-minute panel (30,240 rows, key variables)"],
    ["  1hour_Panel    : 1-hour resampled panel (504 rows)"],
    ["  Daily_Summary  : Daily aggregate statistics (21 rows)"],
    ["  Regime_Summary : Statistics grouped by event regime (4 rows)"],
    ["  LOP_Analysis   : Log LOP deviations and stablecoin FX analysis"],
    ["  Stablecoin_FX  : Stablecoin FX rates and deviations"],
    ["  Data_Inventory : Source and field inventory"],
    [""],
    ["NOTE: Coinbase BTC-USDC public candle API returns identical data to BTC-USD."],
    ["Primary cross-currency LOP analysis relies on Binance.US data."],
]

# ── Write Excel ───────────────────────────────────────────────────────────────
print("Writing Excel workbook (fast mode)...")
with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM:SS") as writer:
    wb = writer.book

    # ── README sheet ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    ws_readme = wb.create_sheet("README", 0)
    ws_readme.sheet_view.showGridLines = False
    title_font = Font(name="Calibri", bold=True, size=14, color="1F3864")
    section_font = Font(name="Calibri", bold=True, size=11, color="1F3864")
    body_font = Font(name="Calibri", size=10)
    section_fill = PatternFill("solid", fgColor="D6E4F0")

    for i, row_data in enumerate(readme_rows, 1):
        text = row_data[0]
        cell = ws_readme.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = title_font
        elif i == 2:
            cell.font = Font(name="Calibri", italic=True, size=12, color="2E75B6")
        elif text and text[0].isupper() and not text.startswith(" ") and not text.startswith("This") and not text.startswith("NOTE"):
            cell.font = section_font
            cell.fill = section_fill
        else:
            cell.font = body_font
    ws_readme.column_dimensions["A"].width = 90

    # ── Write data sheets using pandas ──
    print("  Writing 1min_Panel...")
    df_1m[cols_1m].round(8).to_excel(writer, sheet_name="1min_Panel", index=False)

    print("  Writing 1hour_Panel...")
    cols_1h = [c for c in cols_1m if c in df_1h.columns]
    df_1h[cols_1h].round(8).to_excel(writer, sheet_name="1hour_Panel", index=False)

    print("  Writing Daily_Summary...")
    df_day.round(6).to_excel(writer, sheet_name="Daily_Summary", index=False)

    print("  Writing Regime_Summary...")
    df_regime.round(4).to_excel(writer, sheet_name="Regime_Summary", index=False)

    print("  Writing LOP_Analysis...")
    df_1m[cols_lop].round(8).to_excel(writer, sheet_name="LOP_Analysis", index=False)

    print("  Writing Stablecoin_FX...")
    df_1m[cols_fx].round(8).to_excel(writer, sheet_name="Stablecoin_FX", index=False)

    print("  Writing Data_Inventory...")
    df_inv.to_excel(writer, sheet_name="Data_Inventory", index=False)

    # ── Apply header formatting to all data sheets ──
    print("  Applying formatting...")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    body_font_sm = Font(name="Calibri", size=9)

    regime_fills = {
        "pre_crisis": None,
        "crisis":     PatternFill("solid", fgColor="FFD7D7"),
        "recovery":   PatternFill("solid", fgColor="FFF2CC"),
        "post":       PatternFill("solid", fgColor="E2EFDA"),
    }

    for sheet_name in ["1min_Panel", "1hour_Panel", "LOP_Analysis", "Stablecoin_FX",
                        "Daily_Summary", "Regime_Summary", "Data_Inventory"]:
        ws = wb[sheet_name]
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        # Freeze panes
        ws.freeze_panes = "B2"

        # Auto-width (sample first 100 rows for speed)
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in list(col)[:50]:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 28)

    # Color-code regime rows in 1min_Panel (sample only for speed - do 1h panel fully)
    ws_1h_sheet = wb["1hour_Panel"]
    regime_col_idx = None
    for i, cell in enumerate(ws_1h_sheet[1], 1):
        if cell.value == "regime":
            regime_col_idx = i
            break
    if regime_col_idx:
        ncols = ws_1h_sheet.max_column
        for row_idx in range(2, ws_1h_sheet.max_row + 1):
            regime_val = ws_1h_sheet.cell(row=row_idx, column=regime_col_idx).value
            fill = regime_fills.get(regime_val)
            if fill:
                for col_idx in range(1, ncols + 1):
                    ws_1h_sheet.cell(row=row_idx, column=col_idx).fill = fill

    # Color regime summary
    ws_reg = wb["Regime_Summary"]
    for row_idx in range(2, ws_reg.max_row + 1):
        regime_val = ws_reg.cell(row=row_idx, column=1).value
        fill = regime_fills.get(regime_val)
        if fill:
            for col_idx in range(1, ws_reg.max_column + 1):
                ws_reg.cell(row=row_idx, column=col_idx).fill = fill

print(f"\nSaving workbook to: {EXCEL_PATH}")

print(f"✓ Excel workbook saved: {EXCEL_PATH}")
print(f"  File size: {os.path.getsize(EXCEL_PATH) / 1024 / 1024:.1f} MB")
